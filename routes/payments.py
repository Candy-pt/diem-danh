from flask import Blueprint, render_template, redirect, url_for, flash, request, jsonify, send_file
from flask_login import login_required, current_user
from models import db, Payment, Payroll, Employee
from datetime import datetime, date, timedelta
from sqlalchemy import and_
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO

bp = Blueprint('payments', __name__, url_prefix='/payments')

@bp.route('/')
@login_required
def index():
    from datetime import datetime
    
    # Get current month/year
    now = datetime.now()
    current_month = now.month
    current_year = now.year
    
    # Get all payments
    payments = Payment.query.join(Employee).join(Payroll).order_by(Payment.payment_date.desc()).all()
    
    # Calculate statistics for current month
    current_month_payments = Payment.query.join(Employee).join(Payroll).filter(
        db.extract('month', Payment.payment_date) == current_month,
        db.extract('year', Payment.payment_date) == current_year
    ).all()
    
    total_monthly_payments = sum(p.amount for p in current_month_payments)
    paid_count = len([p for p in current_month_payments if p.status == 'completed'])
    pending_count = len([p for p in current_month_payments if p.status == 'pending'])
    failed_count = len([p for p in current_month_payments if p.status == 'failed'])
    
    return render_template('payments/index.html', 
                         payments=payments,
                         total_monthly_payments=total_monthly_payments,
                         paid_count=paid_count,
                         pending_count=pending_count,
                         failed_count=failed_count,
                         current_month=current_month)

@bp.route('/create', methods=['GET', 'POST'])
@login_required
def create():
    if request.method == 'POST':
        try:
            payment = Payment(
                employee_id=request.form.get('employee_id'),
                payroll_id=request.form.get('payroll_id'),
                amount=float(request.form.get('amount')),
                payment_date=datetime.strptime(request.form.get('payment_date'), '%Y-%m-%d').date(),
                payment_method=request.form.get('payment_method'),
                reference_number=request.form.get('reference_number'),
                status='pending',  # Default status
                notes=request.form.get('notes')
            )
            
            db.session.add(payment)
            db.session.commit()
            
            flash('Payment created successfully!', 'success')
            return redirect(url_for('payments.index'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error creating payment: {str(e)}', 'error')
    
    employees = Employee.query.filter_by(is_active=True).all()
    return render_template('payments/create.html', employees=employees)

@bp.route('/<int:id>')
@login_required
def show(id):
    payment = Payment.query.get_or_404(id)
    return render_template('payments/show.html', payment=payment)

@bp.route('/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit(id):
    payment = Payment.query.get_or_404(id)
    
    if request.method == 'POST':
        try:
            payment.amount = float(request.form.get('amount'))
            payment.payment_date = datetime.strptime(request.form.get('payment_date'), '%Y-%m-%d').date()
            payment.payment_method = request.form.get('payment_method')
            payment.reference_number = request.form.get('reference_number')
            payment.status = request.form.get('status')
            payment.notes = request.form.get('notes')
            
            db.session.commit()
            flash('Payment updated successfully!', 'success')
            return redirect(url_for('payments.show', id=payment.id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating payment: {str(e)}', 'error')
    
    employees = Employee.query.filter_by(is_active=True).all()
    payrolls = Payroll.query.all()
    return render_template('payments/edit.html', payment=payment, employees=employees, payrolls=payrolls)

@bp.route('/<int:id>/mark-paid', methods=['POST'])
@login_required
def mark_as_paid(id):
    payment = Payment.query.get_or_404(id)
    
    try:
        payment.status = 'completed'
        db.session.commit()
        
        if request.headers.get('Content-Type') == 'application/json':
            return jsonify({'success': True, 'message': 'Payment marked as completed!'})
        else:
            flash('Payment marked as completed!', 'success')
            return redirect(url_for('payments.index'))
            
    except Exception as e:
        db.session.rollback()
        error_msg = f'Error updating payment: {str(e)}'
        
        if request.headers.get('Content-Type') == 'application/json':
            return jsonify({'success': False, 'message': error_msg})
        else:
            flash(error_msg, 'error')
            return redirect(url_for('payments.index'))

@bp.route('/<int:id>/delete', methods=['POST', 'DELETE'])
@login_required
def delete(id):
    payment = Payment.query.get_or_404(id)
    
    try:
        db.session.delete(payment)
        db.session.commit()
        
        if request.method == 'DELETE':
            return jsonify({'success': True, 'message': 'Payment deleted successfully!'})
        else:
            flash('Payment deleted successfully!', 'success')
            return redirect(url_for('payments.index'))
            
    except Exception as e:
        db.session.rollback()
        error_msg = f'Error deleting payment: {str(e)}'
        
        if request.method == 'DELETE':
            return jsonify({'success': False, 'message': error_msg})
        else:
            flash(error_msg, 'error')
            return redirect(url_for('payments.index'))

# API routes for AJAX calls
@bp.route('/api/employee/<int:employee_id>/payrolls')
@login_required
def get_employee_payrolls(employee_id):
    payrolls = Payroll.query.filter_by(employee_id=employee_id).all()
    return jsonify([{
        'id': p.id,
        'month': p.month,
        'year': p.year,
        'total_salary': p.total_salary,
        'status': p.status
    } for p in payrolls])

@bp.route('/api/payroll/<int:payroll_id>')
@login_required
def get_payroll_details(payroll_id):
    payroll = Payroll.query.get_or_404(payroll_id)
    return jsonify({
        'id': payroll.id,
        'month': payroll.month,
        'year': payroll.year,
        'basic_salary': payroll.basic_salary,
        'allowance': payroll.allowance,
        'overtime_pay': payroll.overtime_pay,
        'total_salary': payroll.total_salary,
        'status': payroll.status
    })

@bp.route('/bulk-payment', methods=['GET', 'POST'])
@login_required
def bulk_payment():
    if request.method == 'POST':
        payroll_ids = request.form.getlist('payroll_ids')
        payment_date = datetime.strptime(request.form.get('payment_date'), '%Y-%m-%d').date()
        payment_method = request.form.get('payment_method')
        
        if not payroll_ids:
            flash('Please select at least one payroll!', 'error')
            return redirect(url_for('payments.bulk_payment'))
        
        success_count = 0
        error_count = 0
        
        for payroll_id in payroll_ids:
            try:
                payroll = Payroll.query.get(payroll_id)
                if not payroll:
                    continue
                
                # Check if payment already exists
                existing_payment = Payment.query.filter_by(payroll_id=payroll_id).first()
                if existing_payment:
                    flash(f'Payment for payroll {payroll_id} already exists', 'warning')
                    continue
                
                payment = Payment(
                    employee_id=payroll.employee_id,
                    payroll_id=payroll_id,
                    amount=payroll.total_salary,
                    payment_date=payment_date,
                    payment_method=payment_method,
                    status='completed'
                )
                
                # Update payroll status
                payroll.status = 'paid'
                
                db.session.add(payment)
                success_count += 1
                
            except Exception as e:
                error_count += 1
                flash(f'Error processing payment for payroll {payroll_id}: {str(e)}', 'error')
        
        db.session.commit()
        
        if success_count > 0:
            flash(f'Successfully processed {success_count} payments!', 'success')
        if error_count > 0:
            flash(f'Failed to process {error_count} payments!', 'error')
        
        return redirect(url_for('payments.index'))
    
    # Get pending payrolls (approved or calculated status)
    pending_payrolls = Payroll.query.filter(
        Payroll.status.in_(['approved', 'calculated'])
    ).join(Employee).order_by(Payroll.month.desc(), Payroll.year.desc()).all()
    
    return render_template('payments/bulk_payment.html', pending_payrolls=pending_payrolls)

@bp.route('/export-excel')
@login_required
def export_excel():
    start_date = request.args.get('start_date', (date.today() - timedelta(days=30)).strftime('%Y-%m-%d'))
    end_date = request.args.get('end_date', date.today().strftime('%Y-%m-%d'))
    
    payments = Payment.query.join(Employee).join(Payroll).filter(
        and_(
            Payment.payment_date >= datetime.strptime(start_date, '%Y-%m-%d').date(),
            Payment.payment_date <= datetime.strptime(end_date, '%Y-%m-%d').date()
        )
    ).order_by(Payment.payment_date.desc()).all()
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Payments {start_date} to {end_date}"
    
    # Headers
    headers = ['Payment ID', 'Employee ID', 'Employee Name', 'Payroll Period', 'Amount', 
               'Payment Date', 'Payment Method', 'Status', 'Reference Number']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Data
    for row, payment in enumerate(payments, 2):
        ws.cell(row=row, column=1, value=payment.id)
        ws.cell(row=row, column=2, value=payment.employee.employee_id)
        ws.cell(row=row, column=3, value=f"{payment.employee.first_name} {payment.employee.last_name}")
        ws.cell(row=row, column=4, value=f"{payment.payroll.month}/{payment.payroll.year}")
        ws.cell(row=row, column=5, value=payment.amount)
        ws.cell(row=row, column=6, value=payment.payment_date.strftime('%Y-%m-%d'))
        ws.cell(row=row, column=7, value=payment.payment_method)
        ws.cell(row=row, column=8, value=payment.status)
        ws.cell(row=row, column=9, value=payment.reference_number or '')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return send_file(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'payments_{start_date}_to_{end_date}.xlsx'
    )

@bp.route('/report')
@login_required
def report():
    month = request.args.get('month', datetime.now().month)
    year = request.args.get('year', datetime.now().year)
    
    payments = Payment.query.join(Employee).join(Payroll).filter(
        and_(
            Payroll.month == int(month),
            Payroll.year == int(year)
        )
    ).all()
    
    # Calculate summary statistics
    total_payments = len(payments)
    total_amount = sum([p.amount for p in payments])
    completed_payments = len([p for p in payments if p.status == 'completed'])
    pending_payments = len([p for p in payments if p.status == 'pending'])
    
    # Payment method breakdown
    payment_methods = {}
    for payment in payments:
        method = payment.payment_method
        if method not in payment_methods:
            payment_methods[method] = {'count': 0, 'amount': 0}
        payment_methods[method]['count'] += 1
        payment_methods[method]['amount'] += payment.amount
    
    return render_template('payments/report.html',
                         payments=payments,
                         month=month,
                         year=year,
                         total_payments=total_payments,
                         total_amount=total_amount,
                         completed_payments=completed_payments,
                         pending_payments=pending_payments,
                         payment_methods=payment_methods)

@bp.route('/api/payrolls/<int:employee_id>')
@login_required
def api_payrolls(employee_id):
    payrolls = Payroll.query.filter_by(employee_id=employee_id, status='approved').all()
    return jsonify([{
        'id': p.id,
        'period': f"{p.month}/{p.year}",
        'total_salary': p.total_salary,
        'status': p.status
    } for p in payrolls])
