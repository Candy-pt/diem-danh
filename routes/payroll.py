from flask import Blueprint, render_template, redirect, url_for, flash, request, jsonify, send_file
from flask_login import login_required, current_user
from models import db, Payroll, Employee, Attendance
from datetime import datetime, date, timedelta
from sqlalchemy import and_, func
import calendar
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO

bp = Blueprint('payroll', __name__, url_prefix='/payroll')

def calculate_payroll(employee_id, month, year):
    """Calculate payroll for an employee for a specific month"""
    # Get employee
    employee = Employee.query.get(employee_id)
    if not employee:
        return None
    
    # Get month boundaries
    start_date = date(year, month, 1)
    if month == 12:
        end_date = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        end_date = date(year, month + 1, 1) - timedelta(days=1)
    
    # Get attendance records for the month
    attendances = Attendance.query.filter(
        and_(
            Attendance.employee_id == employee_id,
            Attendance.date >= start_date,
            Attendance.date <= end_date
        )
    ).all()
    
    # Calculate working days and hours
    working_days = len([a for a in attendances if a.status == 'present'])
    absent_days = len([a for a in attendances if a.status == 'absent'])
    total_hours = sum([a.total_hours or 0 for a in attendances])
    overtime_hours = sum([a.overtime_hours or 0 for a in attendances])
    
    # Calculate salary components
    daily_salary = employee.salary / 22  # Assuming 22 working days per month
    basic_salary = daily_salary * working_days
    allowance = employee.allowance
    overtime_pay = overtime_hours * (daily_salary / 8) * 1.5  # 1.5x for overtime
    bonus = 0  # Can be configured
    deductions = absent_days * daily_salary
    
    total_salary = basic_salary + allowance + overtime_pay + bonus - deductions
    
    return {
        'basic_salary': round(basic_salary, 2),
        'allowance': allowance,
        'overtime_pay': round(overtime_pay, 2),
        'bonus': bonus,
        'deductions': round(deductions, 2),
        'total_salary': round(total_salary, 2),
        'working_days': working_days,
        'absent_days': absent_days,
        'overtime_hours': overtime_hours
    }

@bp.route('/')
@login_required
def index():
    month = request.args.get('month', datetime.now().month)
    year = request.args.get('year', datetime.now().year)
    
    payrolls = Payroll.query.filter(
        and_(
            Payroll.month == int(month),
            Payroll.year == int(year)
        )
    ).join(Employee).all()
    
    # Calculate summary statistics
    total_monthly_salary = sum([p.total_salary for p in payrolls])
    paid_employees = len([p for p in payrolls if p.status == 'paid'])
    total_overtime = sum([p.overtime_hours for p in payrolls])
    total_allowances = sum([p.allowance for p in payrolls])
    
    return render_template('payroll/index.html', 
                         payrolls=payrolls, 
                         month=month, 
                         year=year,
                         current_month=month,
                         total_monthly_salary=total_monthly_salary,
                         paid_employees=paid_employees,
                         total_overtime=total_overtime,
                         total_allowances=total_allowances)

@bp.route('/generate', methods=['GET', 'POST'])
@login_required
def generate():
    if request.method == 'POST':
        month = int(request.form.get('month'))
        year = int(request.form.get('year'))
        employee_ids = request.form.getlist('employee_ids')
        
        if not employee_ids:
            flash('Please select at least one employee!', 'error')
            return redirect(url_for('payroll.generate'))
        
        success_count = 0
        error_count = 0
        
        for employee_id in employee_ids:
            try:
                # Check if payroll already exists
                existing_payroll = Payroll.query.filter(
                    and_(
                        Payroll.employee_id == employee_id,
                        Payroll.month == month,
                        Payroll.year == year
                    )
                ).first()
                
                if existing_payroll:
                    flash(f'Payroll for employee {employee_id} already exists for {month}/{year}', 'warning')
                    continue
                
                # Calculate payroll
                payroll_data = calculate_payroll(employee_id, month, year)
                if not payroll_data:
                    error_count += 1
                    continue
                
                # Create payroll record
                payroll = Payroll(
                    employee_id=employee_id,
                    month=month,
                    year=year,
                    basic_salary=payroll_data['basic_salary'],
                    allowance=payroll_data['allowance'],
                    overtime_pay=payroll_data['overtime_pay'],
                    bonus=payroll_data['bonus'],
                    deductions=payroll_data['deductions'],
                    total_salary=payroll_data['total_salary'],
                    working_days=payroll_data['working_days'],
                    absent_days=payroll_data['absent_days'],
                    overtime_hours=payroll_data['overtime_hours']
                )
                
                db.session.add(payroll)
                success_count += 1
                
            except Exception as e:
                error_count += 1
                flash(f'Error generating payroll for employee {employee_id}: {str(e)}', 'error')
        
        db.session.commit()
        
        if success_count > 0:
            flash(f'Successfully generated {success_count} payroll records!', 'success')
        if error_count > 0:
            flash(f'Failed to generate {error_count} payroll records!', 'error')
        
        return redirect(url_for('payroll.index', month=month, year=year))
    
    employees = Employee.query.filter_by(is_active=True).all()
    return render_template('payroll/generate.html', employees=employees)

@bp.route('/<int:id>')
@login_required
def show(id):
    payroll = Payroll.query.get_or_404(id)
    return render_template('payroll/show.html', payroll=payroll)

@bp.route('/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit(id):
    payroll = Payroll.query.get_or_404(id)
    
    if request.method == 'POST':
        try:
            payroll.basic_salary = float(request.form.get('basic_salary'))
            payroll.allowance = float(request.form.get('allowance'))
            payroll.overtime_pay = float(request.form.get('overtime_pay'))
            payroll.bonus = float(request.form.get('bonus'))
            payroll.deductions = float(request.form.get('deductions'))
            payroll.total_salary = float(request.form.get('total_salary'))
            payroll.status = request.form.get('status')
            payroll.updated_at = datetime.utcnow()
            
            db.session.commit()
            flash('Payroll updated successfully!', 'success')
            return redirect(url_for('payroll.show', id=payroll.id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating payroll: {str(e)}', 'error')
    
    return render_template('payroll/edit.html', payroll=payroll)

@bp.route('/<int:id>/delete', methods=['POST'])
@login_required
def delete(id):
    payroll = Payroll.query.get_or_404(id)
    
    try:
        db.session.delete(payroll)
        db.session.commit()
        flash('Payroll deleted successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting payroll: {str(e)}', 'error')
    
    return redirect(url_for('payroll.index'))

@bp.route('/export_excel')
@login_required
def export_excel():
    month = request.args.get('month', datetime.now().month)
    year = request.args.get('year', datetime.now().year)
    
    payrolls = Payroll.query.filter(
        and_(
            Payroll.month == int(month),
            Payroll.year == int(year)
        )
    ).join(Employee).all()
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Payroll {month}_{year}"
    
    # Headers
    headers = ['Employee ID', 'Name', 'Department', 'Basic Salary', 'Allowance', 
               'Overtime Pay', 'Bonus', 'Deductions', 'Total Salary', 'Working Days', 'Status']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Data
    for row, payroll in enumerate(payrolls, 2):
        ws.cell(row=row, column=1, value=payroll.employee.employee_id)
        ws.cell(row=row, column=2, value=f"{payroll.employee.first_name} {payroll.employee.last_name}")
        ws.cell(row=row, column=3, value=payroll.employee.department.name)
        ws.cell(row=row, column=4, value=payroll.basic_salary)
        ws.cell(row=row, column=5, value=payroll.allowance)
        ws.cell(row=row, column=6, value=payroll.overtime_pay)
        ws.cell(row=row, column=7, value=payroll.bonus)
        ws.cell(row=row, column=8, value=payroll.deductions)
        ws.cell(row=row, column=9, value=payroll.total_salary)
        ws.cell(row=row, column=10, value=payroll.working_days)
        ws.cell(row=row, column=11, value=payroll.status)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return send_file(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'payroll_{month}_{year}.xlsx'
    )

@bp.route('/report')
@login_required
def report():
    month = request.args.get('month', datetime.now().month)
    year = request.args.get('year', datetime.now().year)
    
    payrolls = Payroll.query.filter(
        and_(
            Payroll.month == int(month),
            Payroll.year == int(year)
        )
    ).join(Employee).all()
    
    # Calculate summary statistics
    total_employees = len(payrolls)
    total_salary = sum([p.total_salary for p in payrolls])
    total_basic_salary = sum([p.basic_salary for p in payrolls])
    total_allowance = sum([p.allowance for p in payrolls])
    total_overtime = sum([p.overtime_pay for p in payrolls])
    total_bonus = sum([p.bonus for p in payrolls])
    total_deductions = sum([p.deductions for p in payrolls])
    
    return render_template('payroll/report.html',
                         payrolls=payrolls,
                         month=month,
                         year=year,
                         total_employees=total_employees,
                         total_salary=total_salary,
                         total_basic_salary=total_basic_salary,
                         total_allowance=total_allowance,
                         total_overtime=total_overtime,
                         total_bonus=total_bonus,
                         total_deductions=total_deductions)
